import numpy as np
from shapely import Polygon
import xlwings as xw
import pandas as pd
import time
from openpyxl import load_workbook
from shapely.geometry import box, Polygon
from shapely import affinity
from rasterio import features
from affine import Affine
from src.auto_plano.zona import Zona

# Tus datos de entrada
aforo = {
    "aforoInicial": 20,
    "aulaInicial": 2,
    "aforoPrimaria": 20,
    "aulaPrimaria": 6,
    "aforoSecundaria": 10,
    "aulaSecundaria": 5
}

archivo = "plantilla.xlsx"

def procesar_excel_real(datos, ruta_archivo):
    # 1. Iniciar Excel (visible=False para que no salte la ventana)
    app = xw.App(visible=False)
    try:
        wb = app.books.open(ruta_archivo)
        
        # --- HOJA INICIAL ---
        ws_inicial = wb.sheets["INICIAL"]
        ws_inicial.range("D5").value = datos["aforoInicial"]
        ws_inicial.range("D6").value = datos["aforoInicial"]
        ws_inicial.range("E3").value = datos["aforoInicial"]
        ws_inicial.range("C2").value = datos["aulaInicial"]
        
        # --- HOJA PRIM ---
        ws_primaria = wb.sheets["PRIM"]
        ws_primaria.range("E3").value = datos["aforoPrimaria"]
        # Llenamos el rango D5:D10 de un solo golpe
        ws_primaria.range("D5:D10").value = [[datos["aforoPrimaria"]]] * 6
        
        # --- HOJA SEC ---
        ws_sec = wb.sheets["SEC"]
        ws_sec.range("E3").value = datos["aforoSecundaria"]
        ws_sec.range("D5:D9").value = [[datos["aforoSecundaria"]]] * 5

        # 2. FORZAR CÁLCULO (Esto es lo que esperabas)
        # Excel suele calcular automáticamente, pero esto asegura que todo esté listo
        wb.app.calculate()
        time.sleep(2) # Un segundo opcional por si el Excel es muy pesado

        # 3. EXTRAER HOJA CALCULOS A DATAFRAME
        ws_calculos = wb.sheets["CALCULOS"]
        
        # xlwings permite convertir un rango directamente a DataFrame de Pandas
        # df_calculos = ws_calculos.used_range.options(pd.DataFrame, index=False, header=True).value
        
        # Limpieza rápida
        # df_calculos.dropna(how="all", inplace=True)
        
        print("✅ Excel procesado y cálculos actualizados con éxito.")
        
        # Guardar si quieres conservar los cambios en el archivo
        wb.save()
        # return df_calculos

    except Exception as e:
        print(f"❌ Error: {e}")
        return None
    finally:
        # 🔒 SIEMPRE CERRAR PARA NO DEJAR PROCESOS COLGADOS
        wb.close()
        app.quit()

# Ejecución
# procesar_excel_real(aforo, archivo)
    
def extraer_df_calculos(ruta_archivo, nombre_hoja="CALCULOS"):
    """
    Lee una hoja específica de un Excel cargando únicamente los valores 
    calculados por las fórmulas y lo convierte en un DataFrame limpio.
    """
    try:
        # 1. Abrir Excel leyendo SOLO los valores calculados (caché de fórmulas)
        wb = load_workbook(ruta_archivo, data_only=True)
        
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            print(f"❌ Error: La hoja '{nombre_hoja}' no existe en el archivo.")
            return None

        ws = wb[nombre_hoja]

        # 2. Convertir a DataFrame
        data = ws.values
        try:
            columns = next(data)
            df_calculos = pd.DataFrame(data, columns=columns)
        except StopIteration:
            wb.close()
            print(f"⚠️ La hoja '{nombre_hoja}' está vacía.")
            return pd.DataFrame()

        # 3. Limpieza de datos
        # Eliminar filas donde TODO sea NaN
        df_calculos.dropna(how="all", inplace=True)
        
        # Eliminar columnas donde TODO sea NaN (como los 'None' al final que vimos antes)
        df_calculos.dropna(axis=1, how="all", inplace=True)
        
        # Opcional: Eliminar columnas que no tengan nombre (None)
        df_calculos = df_calculos.loc[:, df_calculos.columns.notnull()]
        
        # Resetear índice para que sea correlativo tras borrar filas vacías
        df_calculos.reset_index(drop=True, inplace=True)

        # 4. CERRAR EXCEL
        wb.close()
        
        print(f"✅ Extracción de '{nombre_hoja}' completada. {df_calculos.shape[0]} filas procesadas.")
        return df_calculos

    except Exception as e:
        print(f"❌ Ocurrió un error al procesar el Excel: {e}")
        return None


# df_resultados = extraer_df_calculos(archivo)

# print(df_resultados)

def procesar_geometria_utm(utm_coords):
    """
    Normaliza coordenadas UTM al origen (0,0), crea un objeto Polygon
    y extrae sus propiedades principales.
    """
    if not utm_coords:
        raise ValueError("La lista de coordenadas UTM está vacía.")

    # Convertimos a matriz de NumPy para operaciones rápidas
    coords_matrix = np.array(utm_coords)

    # 1. Normalización (Traslación al origen)
    # Guardamos x0 y y0 para poder "des-normalizar" si es necesario después
    x0, y0 = coords_matrix[:, 0].min(), coords_matrix[:, 1].min()
    
    x_norm = coords_matrix[:, 0] - x0
    y_norm = coords_matrix[:, 1] - y0
    
    # Creamos la lista de coordenadas normalizadas (x, y)
    coords_normalizadas = list(zip(x_norm, y_norm))

    # 2. Crear el objeto Polygon de Shapely
    poly = Polygon(coords_normalizadas)

    # 3. Extraer Propiedades
    # El centroide también estará en coordenadas normalizadas
    centroide = poly.centroid
    
    # Preparamos el diccionario de retorno con todo lo necesario para el flujo
    resultado = {
        "polygon": poly,
        "coords_normalizadas": coords_normalizadas,
        "offset": (x0, y0),  # Clave para volver a UTM
        "propiedades": {
            "area": poly.area,
            "perimetro": poly.length,
            "centroide_relativo": (centroide.x, centroide.y),
            "bbox": poly.bounds  # (minx, miny, maxx, maxy)
        }
    }

    print(f"✅ Geometría procesada: Área {poly.area:.2f} m² | Perímetro {poly.length:.2f} m")
    return resultado




def maximal_rectangle(matrix):
    if not matrix.any():
        return 0, (0, 0, 0, 0)

    max_area = 0
    max_rect = (0, 0, 0, 0)
    dp = [0] * len(matrix[0])

    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            dp[j] = dp[j] + 1 if matrix[i][j] == 1 else 0

        stack = []
        for j in range(len(dp) + 1):
            while stack and (j == len(dp) or dp[j] < dp[stack[-1]]):
                height = dp[stack.pop()]
                width = j if not stack else j - stack[-1] - 1
                area = height * width
                if area > max_area:
                    max_area = area
                    top_left_j = stack[-1] + 1 if stack else 0
                    top_left_i = i - height + 1
                    max_rect = (top_left_i, top_left_j, height, width)
            stack.append(j)
    return max_area, max_rect

def find_max_rect_for_angle_fast(polygon, angle_deg, cell_size=1.0):
    # 1. Definir el centro para la rotación
    bounds = polygon.bounds
    minx, miny, maxx, maxy = bounds
    origin = ((minx + maxx) / 2, (miny + maxy) / 2)

    # 2. Rotar el POLÍGONO completo (operación única)
    # Aplicamos un buffer negativo pequeño (-0.01) para asegurar que el
    # rectángulo final esté 100% dentro y no sobrepase por errores decimales.
    poly_rotated = affinity.rotate(polygon.buffer(-0.01), -angle_deg, origin=origin)

    # 3. Preparar la rasterización
    rminx, rminy, rmaxx, rmaxy = poly_rotated.bounds
    width = int(np.ceil((rmaxx - rminx) / cell_size))
    height = int(np.ceil((rmaxy - rminy) / cell_size))

    if width <= 0 or height <= 0:
        return None, 0, angle_deg

    # Transformación Affine para mapear coordenadas a la matriz
    transform = Affine.translation(rminx, rminy) * Affine.scale(cell_size, cell_size)

    # RASTERIZACIÓN VECTORIZADA (Sustituye tus bucles for i, for j)
    grid = features.rasterize(
        [poly_rotated],
        out_shape=(height, width),
        transform=transform,
        fill=0,
        default_value=1,
        dtype=np.uint8
    )

    # 4. Usar tu función maximal_rectangle (que ya es eficiente)
    area_px, (i_start, j_start, h_px, w_px) = maximal_rectangle(grid)

    if area_px == 0:
        return None, 0, angle_deg

    # 5. Reconstruir el rectángulo en el plano rotado
    x0 = j_start * cell_size + rminx
    x1 = (j_start + w_px) * cell_size + rminx
    y0 = i_start * cell_size + rminy
    y1 = (i_start + h_px) * cell_size + rminy

    rect_unrotated = box(x0, y0, x1, y1)

    # 6. Rotar el rectángulo de vuelta a la orientación original
    rect_final = affinity.rotate(rect_unrotated, angle_deg, origin=origin)

    # El área real es proporcional al tamaño de celda
    area_m2 = area_px * (cell_size ** 2)

    return rect_final, area_m2, angle_deg


def find_multiple_max_rectangles_optimized(polygon, angles=np.arange(0, 180, 5), cell_size=0.5, max_rects=3):
    """
    Encuentra hasta max_rects rectángulos máximos dentro de un polígono, usando áreas sobrantes
    y procesando solo los sub-polígonos restantes para cada iteración.
    """
    # Inicializamos la lista de rectángulos
    rectangles = []

    # Lista de polígonos sobrantes, inicialmente el polígono completo
    remaining_polygons = [polygon]

    for _ in range(max_rects):
        best_rect, best_area, best_angle = None, 0, 0
        best_index = -1

        # Revisamos todos los sub-polígonos
        for idx, sub_poly in enumerate(remaining_polygons):
            if sub_poly.is_empty:
                continue

            for angle in angles:
                rect, area, _ = find_max_rect_for_angle_fast(sub_poly, angle, cell_size)
                if rect and area > best_area:
                    best_rect, best_area, best_angle = rect, area, angle
                    best_index = idx

        # Si no encontramos ningún rectángulo válido, terminamos
        if best_rect is None or best_area <= 0:
            break

        # Guardamos el rectángulo encontrado
        rectangles.append((best_rect, best_area, best_angle))

        # Restamos el rectángulo del sub-polígono donde fue encontrado
        sub_poly = remaining_polygons[best_index].difference(best_rect)

        # Actualizamos la lista de polígonos sobrantes
        remaining_polygons.pop(best_index)
        if sub_poly.is_empty:
            continue
        elif sub_poly.geom_type == 'Polygon':
            remaining_polygons.append(sub_poly)
        elif sub_poly.geom_type == 'MultiPolygon':
            remaining_polygons.extend(sub_poly.geoms)

    return rectangles


from shapely.affinity import translate, rotate

def procesar_rectangulo_recto_al_origen(rect_shapely, angulo_optimo):
    """
    Procesamiento lógico: Endereza el rectángulo, lo traslada al origen (0,0)
    y retorna los parámetros de transformación para uso computacional.
    """
    if rect_shapely is None:
        return None

    # 🔑 Guardar el centroide original para futuras rotaciones inversas
    origin = rect_shapely.centroid

    # 1. Enderezar la geometría respecto a su centroide
    rect_recto = rotate(rect_shapely, -angulo_optimo, origin=origin)

    # 2. Obtener límites para la traslación al origen
    minx, miny, maxx, maxy = rect_recto.bounds
    
    # 3. Trasladar al origen (0,0)
    # xoff y yoff negativos mueven el valor min hacia el cero
    rect_0 = translate(rect_recto, xoff=-minx, yoff=-miny)

    # 4. Cálculo de dimensiones finales
    width_m = maxx - minx
    height_m = maxy - miny
    area_final = rect_0.area

    # 🔁 DEVOLVER DICCIONARIO TÉCNICO
    return {
        # Geometría normalizada (para el algoritmo de empaquetamiento/distribución)
        "geometria": rect_0,

        # Datos métricos
        "ancho": width_m,
        "alto": height_m,
        "area": area_final,

        # Metadatos para revertir la transformación (Reverse mapping)
        "angulo": angulo_optimo,
        "origin": origin,
        "offset": (minx, miny)
    }
    
from shapely.ops import unary_union

def unir_zonas(zonas):
    return unary_union([z.geometria for z in zonas])

from collections import defaultdict

def agrupar_ambientes_por_zona(zonas):
    """
    Retorna:
    {
        "Zona A": [amb1, amb2, amb3],
        "Zona B": [amb4, amb5]
    }
    """

    resultado = defaultdict(list)

    def recorrer(zona, zona_actual=None):
        # Si esta zona ES una zona → se vuelve la zona actual
        if zona.tipo == "zona":
            zona_actual = zona

        # Si es ambiente → lo asociamos a la zona actual
        if zona.tipo == "ambiente" and zona_actual is not None:
            resultado[zona_actual.nombre or zona_actual.id].append(zona)

        # Recursividad
        for sub in zona.subzonas:
            recorrer(sub, zona_actual)

    for z in zonas:
        recorrer(z)

    return resultado

def rectangulo_por_metros(ancho, alto):
    """
    Crea un rectángulo desde (0,0) con ancho y alto en metros
    """
    if ancho <= 0 or alto <= 0:
        raise ValueError("El ancho y alto deben ser positivos")

    return Polygon([
        (0, 0),
        (ancho, 0),
        (ancho, alto),
        (0, alto)
    ])
    
from shapely.ops import unary_union
from itertools import combinations

def generar_muros_piso(ambientes, grosor=0.25):
    muros = []

    # Muros interiores
    for a, b in combinations(ambientes, 2):
        if a.geometria.touches(b.geometria):
            linea = a.geometria.boundary.intersection(b.geometria.boundary)

            if not linea.is_empty:
                muro = linea.buffer(grosor/2, cap_style=2)
                muros.append(muro)

    # Muros exteriores
    union = unary_union([a.geometria for a in ambientes])
    muro_exterior = union.buffer(grosor) - union
    muros.append(muro_exterior)

    return muros


def procesar_distribucion_principal(df_calculos, datos_finales, exist_2do_cuad):
    """
    Ejecuta la lógica de zonificación, inserción de ambientes por pabellón
    y cálculo de pasillos para el terreno principal.
    """
    
    # 1. Parámetros base desde el DataFrame
    df_temp_sum = df_calculos.loc[df_calculos['Ambientes'] == 'SUM']
    largo_sum = df_temp_sum['Largo'].values[0] if not df_temp_sum.empty else 10

    # 2. Inicialización del Terreno
    terreno = Zona(datos_finales["geometria"], tipo="ambiente")

    # 3. División inicial [Primaria, Pasillo, Centro, Pasillo, Secundaria]
    # Orientación horizontal según tu esquema
    zonas = terreno.dividir([8, 2, "auto", 2, 8], orientacion="horizontal")
    primaria, area_pasillo_prim, zona2, area_pasillo_sec, secundaria = zonas

    # División de la zona central (zona2) en [Inicial, Medio, Admin]
    zona2.dividir([10, "auto", 10], orientacion="vertical")
    
    inicial = zona2.subzonas[0]
    zona5 = zona2.subzonas[1]  # MEDIO
    admin = zona2.subzonas[2]

    # Asignación de nombres para trazabilidad
    primaria.nombre, secundaria.nombre = "PRIMARIA", "SECUNDARIA"
    inicial.nombre, zona5.nombre, admin.nombre = "INICIAL", "MEDIO", "ADMIN"

    # 4. Sub-división del Pabellón MEDIO
    height_zona_centro_superior = largo_sum + 5
    minx, miny, maxx, maxy = zona5.geometria.bounds
    alto_disponible = maxy - miny

    # Verificación de seguridad para evitar crash en división
    if (12 + height_zona_centro_superior) > alto_disponible:
        height_zona_centro_superior = max(0, alto_disponible - 13)

    # División del medio: [Patio/Inferior, Centro/Losas, SUM/Superior]
    zonas_medio = zona5.dividir([12, "auto", height_zona_centro_superior], orientacion="vertical")
    bottom, center, top = zonas_medio
    
    # Sub-división del centro para Losas y Talleres
    zonas_centro = center.dividir([6, "auto", 6], orientacion="horizontal")
    esp_1, centro_real, esp_2 = zonas_centro
    centro_losas, centro_derecha = centro_real.dividir(["auto", 10], orientacion="horizontal")

    # 5. Mapeo de Pabellones para iteración
    mapa_pabellon = {
        "Izquierda": primaria,
        "Derecha": secundaria,
        "Inferior": inicial,
        "Medio": zona5,
        "Superior": admin
    }

    # Categorías de ambientes especiales
    especiales = {
        "losa": ["Losa Deportiva"],
        "patio": ["Patio de Inicial"],
        "taller": ["Taller EPT"],
        "sum": ["SUM"],
        "cocina": ["Cocina Prim - Sec"]
    }

    # 6. Iteración e Inserción de Ambientes
    for _, row in df_calculos.iterrows():
        if row["Unitario"] == 0: continue
        
        zona_destino = mapa_pabellon.get(row["Pabellon"])
        if zona_destino is None: continue

        for _ in range(int(row["Cantidad"])):
            ambiente = Zona(
                rectangulo_por_metros(row["Ancho"], row["Largo"]),
                nivel=0, nombre=row["Ambientes"], tipo="ambiente", grosor_muro=0.25
            )
            ambiente.aplicar_borde_interior(0.5)

            # Lógica de inserción según Pabellón
            pab = row["Pabellon"]
            
            if pab == "Superior":
                zona_destino.insertar_zona(ambiente)

            elif pab in ["Izquierda", "Derecha"]:
                if exist_2do_cuad:
                    # Si hay 2do cuadrante, intentamos insertar sin forzar piso
                    res = zona_destino.insertar_auto(ambiente, autoPiso=False)
                    if res is ambiente: # Si no entró, se va a inicial como contingencia
                        inicial.insertar_zona(ambiente)
                else:
                    zona_destino.insertar_auto(ambiente, autoPiso=True)

            elif pab == "Inferior":
                if not exist_2do_cuad:
                    zona_destino.insertar_zona(ambiente)

            elif pab == "Medio":
                nombre_amb = row["Ambientes"]
                if nombre_amb in especiales["losa"]:
                    # Lógica de losas múltiples (max 3)
                    if centro_losas.obtener_area_libre() >= ambiente.area_m2:
                        centro_losas.insertar_zona(ambiente, modo="horizontal", gap=2.0)
                elif nombre_amb in especiales["patio"]:
                    if not exist_2do_cuad: bottom.insertar_zona(ambiente, modo="horizontal")
                elif nombre_amb in especiales["sum"] or nombre_amb in especiales["cocina"]:
                    top.insertar_zona(ambiente)
                elif nombre_amb in especiales["taller"]:
                    centro_derecha.insertar_zona(ambiente)
            else:
                zona_destino.insertar_auto(ambiente)

    # 7. Finalización: Centrado y Pasillos
    for z in [secundaria, inicial, zona5, admin, top, bottom, centro_derecha, centro_losas]:
        z.centrar_ambientes()

    # Cálculo dinámico de pasillos basado en la ocupación real
    med_prim = primaria.obtener_ocupacion_vertical()
    med_sec = secundaria.obtener_ocupacion_vertical()

    if med_prim > 0:
        p_prim = Zona(rectangulo_por_metros(2, med_prim), nombre="Pasillo primaria", tipo="ambiente")
        area_pasillo_prim.insertar_zona(p_prim)

    if med_sec > 0:
        p_sec = Zona(rectangulo_por_metros(2, med_sec), nombre="Pasillo secundaria", tipo="ambiente")
        area_pasillo_sec.insertar_zona(p_sec)

    # 8. Retorno de objetos para flujo posterior
    return {
        "terreno": terreno,
        "zonas_completas": zonas,
        "primaria": primaria,
        "secundaria": secundaria,
        "inicial": inicial,
        "medio": zona5,
        "admin": admin,
        "medidas": {"primaria": med_prim, "secundaria": med_sec}
    }
    
from shapely.geometry import Polygon
import pandas as pd

def reconstruir_zonas(terrenos):
    """
    Recibe un terreno o una lista de terrenos.
    Devuelve:
        - zonas_reconstruidas (dict)
        - df (DataFrame con la data original)
    """

    # Permitir que sea uno solo o lista
    if not isinstance(terrenos, (list, tuple)):
        terrenos = [terrenos]

    data_total = []

    for terreno in terrenos:
        data = terreno.obtener_geometrias_recursivas()
        data_total.extend(data)

    zonas_reconstruidas = {
        d["id"]: {
            "path": d.get("path"),
            "nombre": d.get("nombre"),
            "geometria": Polygon(d["coords"]),
            "piso": d.get("piso"),
        }
        for d in data_total
    }

    df = pd.DataFrame(data_total)

    return zonas_reconstruidas, df

def local_a_mundo(coords, transformacion):
    poly = Polygon(coords)

    # 1. deshacer traslado al (0,0)
    poly = translate(
        poly,
        xoff=transformacion["offset"][0],
        yoff=transformacion["offset"][1]
    )

    # 2. deshacer rotación
    poly = rotate(
        poly,
        transformacion["angulo"],
        origin=transformacion["origin"]
    )

    return poly

# def procesar_multiple_terrenos(lista_terrenos):
#     from shapely.geometry import Polygon
#     import pandas as pd

#     data_total = []

#     for terreno, best_rect, best_angle in lista_terrenos:

#         # 1️⃣ Obtener data local
#         data = terreno.obtener_geometrias_recursivas()

#         # 2️⃣ Calcular transformación propia
#         datos_finales = procesar_rectangulo_recto_al_origen(best_rect, best_angle)

#         transformacion = {
#             "angulo": datos_finales["angulo"],
#             "origin": datos_finales["origin"],
#             "offset": datos_finales["offset"]
#         }

#         # 3️⃣ Transformar cada geometría
#         for d in data:
#             poly_local = Polygon(d["coords"])
#             d["geometria_mundo"] = local_a_mundo(poly_local, transformacion)

#         data_total.extend(data)

#     df = pd.DataFrame(data_total)
#     return df

from shapely.geometry import Polygon
import pandas as pd
import numpy as np

def procesar_multiple_terrenos(lista_terrenos, utm_coords_terreno):
    data_total = []
    
    # 1️⃣ Validar y agregar Terreno Principal
    try:
        poly_terreno = Polygon(utm_coords_terreno)
        terreno_perimetro = {
            "id": "terreno_principal",
            "nombre": "Límite Terreno",
            "tipo": "perimetro",
            "piso": 0,
            "geometria_mundo": poly_terreno 
        }
        data_total.append(terreno_perimetro)
    except Exception as e:
        print(f"❌ Error creando polígono de terreno: {e}")

    # 2️⃣ Procesar lista de terrenos (pabellones)
    for i, (terreno, best_rect, best_angle) in enumerate(lista_terrenos):
        try:
            # Obtener data local
            data = terreno.obtener_geometrias_recursivas()

            # Calcular transformación
            data_process = procesar_rectangulo_recto_al_origen(best_rect, best_angle)
            
            # 🛡️ DEPURACIÓN: Ver si la transformación tiene NaNs
            if any(np.isnan(v) if isinstance(v, (int, float)) else False for v in data_process.values()):
                print(f"⚠️ Alerta: data_process del terreno {i} contiene NaNs: {data_process}")

            transformacion = {
                "angulo": data_process["angulo"],
                "origin": data_process["origin"],
                "offset": data_process["offset"]
            }

            # 3️⃣ Transformar cada geometría
            for d in data:
                try:
                    poly_local = Polygon(d["coords"])
                    d["geometria_mundo"] = local_a_mundo(poly_local, transformacion)
                except Exception as e:
                    print(f"❌ Error transformando ambiente '{d.get('nombre')}': {e}")
                    d["geometria_mundo"] = None

            data_total.extend(data)
            
        except Exception as e:
            print(f"❌ Error procesando terreno {i}: {e}")

    # 4️⃣ LIMPIEZA FINAL PARA JSON (Crucial para evitar el Error 500)
    df = pd.DataFrame(data_total)

    def preparar_geometria(poly):
        """Convierte Polígono a lista para JSON y maneja errores."""
        try:
            if poly is None or (hasattr(poly, 'is_empty') and poly.is_empty):
                return None
            # Convertimos las coordenadas a lista de listas [[x,y], [x,y]...]
            return [list(p) for p in poly.exterior.coords]
        except:
            return None

    if not df.empty:
        # Convertir objetos Shapely a Listas
        df["geometria_mundo"] = df["geometria_mundo"].apply(preparar_geometria)
        
        # 🛡️ ELIMINAR NaNs: Reemplaza cualquier NaN matemático por None (null en JSON)
        df = df.replace({np.nan: None})
        
        return df.to_dict(orient="records")
    
    return []

import matplotlib.pyplot as plt

def visualizar_distribucion_global(df_global, titulo="Distribución Global de Pabellones", padding=5.0):
    """
    Dibuja todos los pabellones y ambientes en sus coordenadas finales (mundo real).
    Calcula los límites globales para asegurar que todo el proyecto sea visible.
    """
    if df_global.empty:
        print("⚠️ El DataFrame global está vacío. Nada que mostrar.")
        return

    plt.figure(figsize=(12, 12))
    
    # Inicializamos límites con valores extremos
    minx, miny = float("inf"), float("inf")
    maxx, maxy = float("-inf"), float("-inf")

    # Iteramos sobre los elementos para dibujar y calcular el encuadre
    for _, row in df_global.iterrows():
        poly = row.get("geometria_mundo")
        if poly is None or poly.is_empty:
            continue

        # 1. Dibujar el polígono
        x, y = poly.exterior.xy
        plt.plot(x, y, linewidth=1, alpha=0.8, label=row.get("Ambientes", "Sin nombre")[:15])
        plt.fill(x, y, alpha=0.2)

        # 2. Acumular bounds para el encuadre global
        bx, by, Bx, By = poly.bounds
        minx, miny = min(minx, bx), min(miny, by)
        maxx, maxy = max(maxx, Bx), max(maxy, By)

    # 3. Configuración de límites con margen (padding)
    if minx != float("inf"):
        plt.xlim(minx - padding, maxx + padding)
        plt.ylim(miny - padding, maxy + padding)
    
    # 4. Estética de la gráfica
    plt.gca().set_aspect("equal")
    plt.title(titulo, fontsize=14, fontweight="bold")
    plt.grid(True, linestyle="--", alpha=0.3)
    plt.xlabel("Este (UTM)")
    plt.ylabel("Norte (UTM)")
    
    # Opcional: Mostrar leyenda solo si no hay demasiados elementos
    if len(df_global) < 20:
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='small')

    plt.tight_layout()
    plt.show()

# --- Ejemplo de uso ---
# visualizar_distribucion_global(df_final_proyecto)

def procesar_segundo_cuadrante(df_calculos, rects, datos_finales_2do_cuad):
    """
    Procesa la distribución de ambientes en el segundo cuadrante 
    basándose en los cálculos del Excel y la geometría detectada.
    """
    # 1. Validación y Extracción del rectángulo
    if len(rects) < 2:
        raise ValueError("No se encontró un segundo rectángulo para el 2do cuadrante en la lista de 'rects'.")
    
    # Extraemos el rectángulo (best_rect_q2, best_area_q2, best_angle_q2)
    # Según tu flujo, usamos la geometría directamente de los datos del cuadrante
    geometria_q2 = datos_finales_2do_cuad.get("geometria")
    if geometria_q2 is None:
        raise ValueError("No se encontró la geometría en datos_finales_2do_cuad.")

    # 2. Definir terreno y dividir: [8m Aulas, 2m Pasillo, 'auto' Patio]
    terreno_q2 = Zona(geometria_q2, tipo="ambiente")
    zonas_q2 = terreno_q2.dividir([8, 2, "auto"], orientacion="vertical")

    zona_aulas_q2, pasillo_q2, zona_patio_q2 = zonas_q2
    zona_aulas_q2.nombre = "AULAS INICIAL Q2"
    zona_patio_q2.nombre = "PATIO DE INICIAL Q2"

    # 3. Procesar df_calculos para insertar ambientes
    patio_inicial_nombres_q2 = ["Patio de Inicial"]

    for _, row in df_calculos.iterrows():
        # Saltamos si no hay área definida o cantidad es 0
        if row.get("Unitario", 0) == 0 or row.get("Cantidad", 0) == 0:
            continue

        for _ in range(int(row["Cantidad"])):
            # Crear el objeto ambiente con las dimensiones del Excel
            ambiente_q2 = Zona(
                rectangulo_por_metros(row["Ancho"], row["Largo"]),
                nivel=0,
                nombre=row["Ambientes"],
                tipo="ambiente",
                grosor_muro=0.25
            )
            ambiente_q2.aplicar_borde_interior(0.5)

            # Lógica de distribución por Pabellón
            pabellon = row.get("Pabellon")
            
            # CASO 1: Aulas de Inicial (Pabellón Inferior)
            if pabellon == "Inferior":
                zona_aulas_q2.insertar_zona(ambiente_q2)

            # CASO 2: Patio de Inicial (Pabellón Medio)
            elif pabellon == "Medio":
                if row["Ambientes"] in patio_inicial_nombres_q2:
                    print(f"📍 Insertando {row['Ambientes']} en zona PATIO del 2do cuadrante")
                    zona_patio_q2.insertar_zona(ambiente_q2)

    # 4. Ajustes finales: Centrar ambientes dentro de sus contenedores
    zona_aulas_q2.centrar_ambientes()
    zona_patio_q2.centrar_ambientes()

    # # 5. Visualización (Opcional, puedes comentarlo si solo quieres los datos)
    # plotly_visualizar_zonas_por_piso(
    #     zonas_q2,
    #     piso=1,
    #     title="INICIAL: AULAS + PATIO (2DO CUADRANTE)"
    # )

    # Retornamos las zonas procesadas para seguir el flujo (ej. guardarlas o unificarlas)
    return {
        "zonas": zonas_q2,
        "zona_aulas": zona_aulas_q2,
        "zona_patio": zona_patio_q2,
        "pasillo": pasillo_q2
    }
    
import ezdxf
from shapely.geometry import Polygon

def exportar_unico_archivo_cad(df, filename="plano_georeferenciado.dxf"):
    doc = ezdxf.new('R2010')
    msp = doc.modelspace()

    # Configuración de Colores DXF (Índices estándar de AutoCAD)
    COLOR_ROJO = 1
    COLOR_CYAN = 4
    COLOR_BLANCO = 7 
    OFFSET_X = 500 # Espacio entre plantas en el CAD

    # 1. Exportar el PERÍMETRO DEL TERRENO (Capa Global)
    # Lo buscamos en el DF o lo dibujamos una vez como referencia
    df_perimetro = df[df['tipo'] == 'perimetro']
    if not df_perimetro.empty:
        layer_terr = "LIMITE_PROPIEDAD"
        doc.layers.new(name=layer_terr, dxfattribs={'color': COLOR_ROJO, 'linetype': 'DASHED'})
        
        for _, row in df_perimetro.iterrows():
            poly = row['geometria_mundo']
            # Extraer puntos (x, y) del polígono de Shapely
            puntos = list(poly.exterior.coords)
            msp.add_lwpolyline(puntos, close=True, dxfattribs={'layer': layer_terr, 'const_width': 0.1})

    # 2. Exportar AMBIENTES organizados por PISO
    # Agrupamos solo los que no son perímetro
    df_ambientes = df[df['tipo'] == 'ambiente']
    
    for num_piso, df_piso in df_ambientes.groupby('piso'):
        # Desplazamiento para no amontonar los pisos en el CAD
        desplazamiento_actual = (num_piso - 1) * OFFSET_X
        layer_name = f"PISO_{num_piso}"
        
        if layer_name not in doc.layers:
            doc.layers.new(name=layer_name, dxfattribs={'color': COLOR_BLANCO})

        for _, row in df_piso.iterrows():
            poly = row['geometria_mundo']
            if poly is None or poly.is_empty:
                continue

            # Aplicar el desplazamiento a las coordenadas reales del mundo
            puntos_mundo = list(poly.exterior.coords)
            coords_finales = [(x + desplazamiento_actual, y) for x, y in puntos_mundo]

            # Dibujar el ambiente
            msp.add_lwpolyline(
                coords_finales, 
                close=True, 
                dxfattribs={'layer': layer_name}
            )

            # Añadir Texto (Nombre del ambiente)
            centro = poly.centroid
            msp.add_text(
                str(row['nombre']),
                dxfattribs={
                    'layer': layer_name,
                    'height': 0.6,
                    'color': COLOR_CYAN # Texto en cyan para legibilidad
                }
            ).set_placement((centro.x + desplazamiento_actual, centro.y))

        # Etiqueta de título de planta
        # Buscamos un punto base para el título cerca del grupo de polígonos
        base_x = df_piso['geometria_mundo'].apply(lambda p: p.bounds[0]).min()
        base_y = df_piso['geometria_mundo'].apply(lambda p: p.bounds[1]).min()
        
        msp.add_text(
            f"NIVEL {num_piso}",
            dxfattribs={'height': 3.0, 'color': COLOR_BLANCO, 'style': 'OpenSans'}
        ).set_placement((base_x + desplazamiento_actual, base_y - 10))

    doc.saveas(filename)
    print(f"✅ DXF Exportado con éxito: {filename}")