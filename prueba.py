import xlwings as xw
import pandas as pd
import time
from openpyxl import load_workbook

# Tus datos de entrada
aforo = {
    "aforoInicial": 20,
    "aulaInicial": 2,
    "aforoPrimaria": 20,
    "aulaPrimaria": 6,
    "aforoSecundaria": 10,
    "aulaSecundaria": 5
}

archivo = "plantilla.xlsx"

def procesar_excel_real(datos, ruta_archivo):
    # 1. Iniciar Excel (visible=False para que no salte la ventana)
    app = xw.App(visible=False)
    try:
        wb = app.books.open(ruta_archivo)
        
        # --- HOJA INICIAL ---
        ws_inicial = wb.sheets["INICIAL"]
        ws_inicial.range("D5").value = datos["aforoInicial"]
        ws_inicial.range("D6").value = datos["aforoInicial"]
        ws_inicial.range("E3").value = datos["aforoInicial"]
        ws_inicial.range("C2").value = datos["aulaInicial"]
        
        # --- HOJA PRIM ---
        ws_primaria = wb.sheets["PRIM"]
        ws_primaria.range("E3").value = datos["aforoPrimaria"]
        # Llenamos el rango D5:D10 de un solo golpe
        ws_primaria.range("D5:D10").value = [[datos["aforoPrimaria"]]] * 6
        
        # --- HOJA SEC ---
        ws_sec = wb.sheets["SEC"]
        ws_sec.range("E3").value = datos["aforoSecundaria"]
        ws_sec.range("D5:D9").value = [[datos["aforoSecundaria"]]] * 5

        # 2. FORZAR CÁLCULO (Esto es lo que esperabas)
        # Excel suele calcular automáticamente, pero esto asegura que todo esté listo
        wb.app.calculate()
        time.sleep(2) # Un segundo opcional por si el Excel es muy pesado

        # 3. EXTRAER HOJA CALCULOS A DATAFRAME
        ws_calculos = wb.sheets["CALCULOS"]
        
        # xlwings permite convertir un rango directamente a DataFrame de Pandas
        # df_calculos = ws_calculos.used_range.options(pd.DataFrame, index=False, header=True).value
        
        # Limpieza rápida
        # df_calculos.dropna(how="all", inplace=True)
        
        print("✅ Excel procesado y cálculos actualizados con éxito.")
        
        # Guardar si quieres conservar los cambios en el archivo
        wb.save()
        # return df_calculos

    except Exception as e:
        print(f"❌ Error: {e}")
        return None
    finally:
        # 🔒 SIEMPRE CERRAR PARA NO DEJAR PROCESOS COLGADOS
        wb.close()
        app.quit()

# Ejecución
procesar_excel_real(aforo, archivo)
    
def extraer_df_calculos(ruta_archivo, nombre_hoja="CALCULOS"):
    """
    Lee una hoja específica de un Excel cargando únicamente los valores 
    calculados por las fórmulas y lo convierte en un DataFrame limpio.
    """
    try:
        # 1. Abrir Excel leyendo SOLO los valores calculados (caché de fórmulas)
        wb = load_workbook(ruta_archivo, data_only=True)
        
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            print(f"❌ Error: La hoja '{nombre_hoja}' no existe en el archivo.")
            return None

        ws = wb[nombre_hoja]

        # 2. Convertir a DataFrame
        data = ws.values
        try:
            columns = next(data)
            df_calculos = pd.DataFrame(data, columns=columns)
        except StopIteration:
            wb.close()
            print(f"⚠️ La hoja '{nombre_hoja}' está vacía.")
            return pd.DataFrame()

        # 3. Limpieza de datos
        # Eliminar filas donde TODO sea NaN
        df_calculos.dropna(how="all", inplace=True)
        
        # Eliminar columnas donde TODO sea NaN (como los 'None' al final que vimos antes)
        df_calculos.dropna(axis=1, how="all", inplace=True)
        
        # Opcional: Eliminar columnas que no tengan nombre (None)
        df_calculos = df_calculos.loc[:, df_calculos.columns.notnull()]
        
        # Resetear índice para que sea correlativo tras borrar filas vacías
        df_calculos.reset_index(drop=True, inplace=True)

        # 4. CERRAR EXCEL
        wb.close()
        
        print(f"✅ Extracción de '{nombre_hoja}' completada. {df_calculos.shape[0]} filas procesadas.")
        return df_calculos

    except Exception as e:
        print(f"❌ Ocurrió un error al procesar el Excel: {e}")
        return None


df_resultados = extraer_df_calculos(archivo)

print(df_resultados)